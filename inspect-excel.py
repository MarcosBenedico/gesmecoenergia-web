#!/usr/bin/env python3
import openpyxl
import sys

try:
    wb = openpyxl.load_workbook(sys.argv[1])

    print("\n" + "="*70)
    print("ESTRUCTURA DEL EXCEL")
    print("="*70)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📄 HOJA: {sheet_name}")
        print("-"*70)

        max_row = 0
        for row in ws.iter_rows():
            if any(cell.value for cell in row):
                max_row = row[0].row

        print(f"Total filas con datos: {max_row}")
        print("\nContenido:")

        for row_num in range(1, min(max_row + 1, 100)):
            row_data = []
            for col_num in range(1, 10):
                cell = ws.cell(row_num, col_num)
                if cell.value:
                    row_data.append(f"Col{col_num}:'{cell.value}'")

            if row_data:
                print(f"Fila {row_num:2d}: {' | '.join(row_data)}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
